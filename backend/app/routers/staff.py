import secrets
import string
from datetime import date as date_type
from io import BytesIO
from typing import Annotated, Optional
from uuid import UUID

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import select, or_, func
from sqlalchemy.orm import selectinload
from pwdlib import PasswordHash
from pwdlib.hashers.argon2 import Argon2Hasher

from app.dependencies import CurrentUser, CurrentSchool, DB, require_roles
from app.models.staff import Staff, StaffQualification, StaffPromotion
from app.models.user import User
from app.models.ges_rank import GESRank
from app.schemas.staff import (
    StaffCreate, StaffUpdate, StaffSummary, StaffResponse,
    InviteStaff, InviteResponse,
    QualificationCreate, QualificationResponse,
    PromotionCreate, PromotionResponse,
    GESRankItem,
)

router = APIRouter()
_pwd = PasswordHash((Argon2Hasher(),))


def _hash(pw: str) -> str:
    return _pwd.hash(pw)


def _gen_password(length: int = 12) -> str:
    chars = string.ascii_letters + string.digits
    return "".join(secrets.choice(chars) for _ in range(length))


def _can_manage(current_user: User, staff_id: UUID) -> bool:
    """True when the user is managing their own record or has elevated role."""
    if current_user.role in ("school_admin", "headteacher"):
        return True
    return current_user.staff_id == staff_id


def _load_full(q):
    return q.options(
        selectinload(Staff.user),
        selectinload(Staff.qualifications),
        selectinload(Staff.promotions),
    )


# ── GES rank catalogue ────────────────────────────────────────────────────
@router.get("/ges-ranks", response_model=list[GESRankItem])
async def list_ges_ranks(db: DB, _: CurrentUser):
    result = await db.execute(
        select(GESRank).order_by(GESRank.category, GESRank.order)
    )
    return result.scalars().all()


# ── List staff ────────────────────────────────────────────────────────────
@router.get("/", response_model=list[StaffSummary])
async def list_staff(
    db: DB,
    current_user: CurrentUser,
    school: CurrentSchool,
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    role:   Optional[str] = Query(None),
    skip:   int = Query(0,  ge=0),
    limit:  int = Query(50, ge=1, le=500),
):
    q = (
        select(Staff)
        .where(Staff.school_id == school.id)
        .options(
            selectinload(Staff.user),
            selectinload(Staff.promotions),   # needed for current_rank property
        )
    )
    if search:
        term = f"%{search.lower()}%"
        q = q.where(or_(
            func.lower(Staff.first_name).like(term),
            func.lower(Staff.last_name).like(term),
            func.lower(Staff.staff_number).like(term),
            func.lower(Staff.specialization).like(term),
            func.lower(Staff.phone).like(term),
        ))
    if status:
        q = q.where(Staff.status == status)
    if role:
        q = q.join(User, User.staff_id == Staff.id).where(User.role == role)

    q = q.order_by(Staff.last_name, Staff.first_name).offset(skip).limit(limit)
    result = await db.execute(q)
    return result.scalars().all()


# ── Get one ───────────────────────────────────────────────────────────────
@router.get("/{staff_id}", response_model=StaffResponse)
async def get_staff(
    staff_id: UUID,
    db: DB,
    current_user: CurrentUser,
    school: CurrentSchool,
):
    result = await db.execute(
        _load_full(
            select(Staff).where(Staff.id == staff_id, Staff.school_id == school.id)
        )
    )
    staff = result.scalar_one_or_none()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")
    return staff


# ── Create ────────────────────────────────────────────────────────────────
@router.post("/", response_model=InviteResponse, status_code=201)
async def create_staff(
    body: StaffCreate,
    db: DB,
    _: Annotated[User, Depends(require_roles("school_admin", "headteacher"))],
    school: CurrentSchool,
):
    if body.email:
        exists = await db.execute(select(User).where(User.email == body.email.lower()))
        if exists.scalar_one_or_none():
            raise HTTPException(status_code=409, detail="Email already in use")

    staff = Staff(school_id=school.id, **body.model_dump(exclude={"email", "password", "role"}))
    db.add(staff)
    await db.flush()

    temp_password: str | None = None
    if body.email:
        raw_password = body.password or _gen_password()
        temp_password = raw_password if not body.password else None
        db.add(User(
            school_id=school.id,
            email=body.email.lower(),
            password_hash=_hash(raw_password),
            role=body.role,
            staff_id=staff.id,
        ))

    await db.commit()
    result = await db.execute(_load_full(select(Staff).where(Staff.id == staff.id)))
    refreshed = result.scalar_one()
    return InviteResponse(**StaffResponse.model_validate(refreshed).model_dump(), temp_password=temp_password)


# ── Update ────────────────────────────────────────────────────────────────
@router.patch("/{staff_id}", response_model=StaffResponse)
async def update_staff(
    staff_id: UUID,
    body: StaffUpdate,
    db: DB,
    _: Annotated[User, Depends(require_roles("school_admin", "headteacher"))],
    school: CurrentSchool,
):
    result = await db.execute(
        select(Staff).where(Staff.id == staff_id, Staff.school_id == school.id)
    )
    staff = result.scalar_one_or_none()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")

    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(staff, field, value)

    await db.commit()
    result = await db.execute(_load_full(select(Staff).where(Staff.id == staff_id)))
    return result.scalar_one()


# ── Deactivate (soft delete) ──────────────────────────────────────────────
@router.delete("/{staff_id}", status_code=204)
async def deactivate_staff(
    staff_id: UUID,
    db: DB,
    _: Annotated[User, Depends(require_roles("school_admin"))],
    school: CurrentSchool,
):
    result = await db.execute(
        select(Staff)
        .where(Staff.id == staff_id, Staff.school_id == school.id)
        .options(selectinload(Staff.user))
    )
    staff = result.scalar_one_or_none()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")
    staff.status = "retired"
    if staff.user:
        staff.user.is_active = False
    await db.commit()


# ── Invite / reset credentials ────────────────────────────────────────────
@router.post("/{staff_id}/invite", response_model=InviteResponse)
async def invite_staff(
    staff_id: UUID,
    body: InviteStaff,
    db: DB,
    _: Annotated[User, Depends(require_roles("school_admin", "headteacher"))],
    school: CurrentSchool,
):
    result = await db.execute(
        select(Staff)
        .where(Staff.id == staff_id, Staff.school_id == school.id)
        .options(selectinload(Staff.user))
    )
    staff = result.scalar_one_or_none()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")

    raw_password = body.password or _gen_password()

    if staff.user:
        staff.user.email = body.email.lower()
        staff.user.role = body.role
        staff.user.password_hash = _hash(raw_password)
        staff.user.is_active = True
    else:
        exists = await db.execute(select(User).where(User.email == body.email.lower()))
        if exists.scalar_one_or_none():
            raise HTTPException(status_code=409, detail="Email already in use")
        db.add(User(
            school_id=school.id,
            email=body.email.lower(),
            password_hash=_hash(raw_password),
            role=body.role,
            staff_id=staff.id,
        ))

    await db.commit()
    result = await db.execute(_load_full(select(Staff).where(Staff.id == staff_id)))
    refreshed = result.scalar_one()

    return InviteResponse(**StaffResponse.model_validate(refreshed).model_dump(), temp_password=raw_password)


# ── Toggle login ──────────────────────────────────────────────────────────
@router.post("/{staff_id}/toggle-account", response_model=StaffResponse)
async def toggle_account(
    staff_id: UUID,
    db: DB,
    _: Annotated[User, Depends(require_roles("school_admin", "headteacher"))],
    school: CurrentSchool,
):
    result = await db.execute(
        select(Staff)
        .where(Staff.id == staff_id, Staff.school_id == school.id)
        .options(selectinload(Staff.user))
    )
    staff = result.scalar_one_or_none()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")
    if not staff.user:
        raise HTTPException(status_code=400, detail="No login account linked")
    staff.user.is_active = not staff.user.is_active
    await db.commit()
    result = await db.execute(_load_full(select(Staff).where(Staff.id == staff_id)))
    return result.scalar_one()


# ── Bulk upload ───────────────────────────────────────────────────────────

@router.get("/upload/template")
async def download_staff_template(
    _: CurrentUser,
    school: CurrentSchool,
    db: DB,
):
    """School-branded Excel template for bulk staff import."""
    ranks_result = await db.execute(
        select(GESRank).order_by(GESRank.category, GESRank.order)
    )
    ges_ranks = ranks_result.scalars().all()

    wb = openpyxl.Workbook()

    # ── Hidden _data sheet (rank lookup) ─────────────────────────────
    ws_data = wb.create_sheet("_data")
    ws_data.sheet_state = "hidden"
    ws_data["A1"] = "rank_name"
    for i, r in enumerate(ges_ranks, 2):
        ws_data.cell(row=i, column=1, value=r.name)
    num_ranks = len(ges_ranks)

    # ── Main sheet ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Staff"

    school_color = (school.accent_color or "#1a6b3c").lstrip("#").upper()
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", fgColor=school_color)
    required_fill = PatternFill("solid", fgColor="E8F5E9")
    optional_fill = PatternFill("solid", fgColor="F5F5F5")
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Title row
    ws.merge_cells("A1:L1")
    ws["A1"] = f"{school.name}  |  Staff Import"
    ws["A1"].font = Font(bold=True, size=13, color=school_color)
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 30

    # Instructions row
    ws.merge_cells("A2:L2")
    ws["A2"] = (
        "Fill from row 5. Green = required. "
        "Date format: YYYY-MM-DD. Phone: include leading zero e.g. 0244123456. "
        "Status: active | on_leave | transferred | retired (default: active)."
    )
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["A2"].alignment = center_align
    ws.row_dimensions[2].height = 25

    # (header, width, required, example)
    columns = [
        ("staff_number",   15, False, "TCH001"),
        ("first_name",     15, True,  "Kwame"),
        ("middle_name",    15, False, "Asante"),
        ("last_name",      15, True,  "Mensah"),
        ("title",          10, False, "Mr"),
        ("gender",         10, False, "Male"),
        ("date_of_birth",  15, False, "1985-06-20"),
        ("phone",          15, False, "0244123456"),
        ("date_joined",    15, False, "2010-09-01"),
        ("license_number", 18, False, "TCH-00001"),
        ("specialization", 22, False, "Mathematics, Physics"),
        ("status",         12, False, "active"),
    ]

    # Header row (row 3)
    for col_idx, (header, width, required, _) in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 22

    # Example row (row 4)
    for col_idx, (_, _, required, example) in enumerate(columns, 1):
        cell = ws.cell(row=4, column=col_idx, value=example)
        cell.fill = required_fill if required else optional_fill
        cell.font = Font(italic=True, color="888888", size=10)

    # Phone column (H = col 8) as text
    for row in range(4, 10001):
        ws[f"H{row}"].number_format = "@"

    # Data validations
    gender_dv = DataValidation(
        type="list", formula1='"Male,Female"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="Invalid gender", error="Select Male or Female",
    )
    ws.add_data_validation(gender_dv)
    gender_dv.sqref = "F5:F10000"

    title_dv = DataValidation(
        type="list", formula1='"Mr,Mrs,Ms,Dr,Prof,Rev"',
        allow_blank=True,
    )
    ws.add_data_validation(title_dv)
    title_dv.sqref = "E5:E10000"

    status_dv = DataValidation(
        type="list", formula1='"active,on_leave,transferred,retired"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.sqref = "L5:L10000"

    ws.freeze_panes = "A5"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{school.slug}_staff_import.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/upload")
async def bulk_upload_staff(
    _: Annotated[User, Depends(require_roles("school_admin", "headteacher"))],
    school: CurrentSchool,
    db: DB,
    file: UploadFile = File(...),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "File must be an Excel file (.xlsx or .xls)")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "Could not read Excel file. Please use the downloaded template.")

    HEADER_ROW = 3
    DATA_START  = 5

    headers = [str(c.value).strip() for c in ws[HEADER_ROW] if c.value]
    required_headers = {"first_name", "last_name"}
    missing = required_headers - set(headers)
    if missing:
        raise HTTPException(
            400,
            f"Missing required columns: {missing}. Please use the downloaded template.",
        )

    # Pre-load existing staff numbers for duplicate detection
    existing_result = await db.execute(
        select(Staff.staff_number).where(
            Staff.school_id == school.id,
            Staff.staff_number.isnot(None),
        )
    )
    existing_numbers = {row[0] for row in existing_result.all()}

    def parse_date(val, field_name: str, row_errors: list):
        if not val:
            return None
        try:
            if isinstance(val, date_type):
                return val
            return date_type.fromisoformat(str(val).strip())
        except ValueError:
            row_errors.append(f"{field_name} must be YYYY-MM-DD, got: {val!r}")
            return None

    imported = 0
    skipped  = 0
    errors   = []
    seen_numbers: set[str] = set()

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=DATA_START, values_only=True), start=DATA_START
    ):
        if not any(row):
            continue

        row_data: dict[str, object] = {}
        for i, header in enumerate(headers):
            row_data[header] = row[i] if i < len(row) else None

        row_errors: list[str] = []

        first_name = str(row_data.get("first_name", "") or "").strip()
        last_name  = str(row_data.get("last_name", "") or "").strip()
        staff_number = str(row_data.get("staff_number", "") or "").strip() or None

        if not first_name:
            row_errors.append("first_name is required")
        if not last_name:
            row_errors.append("last_name is required")

        if staff_number:
            if staff_number in seen_numbers:
                row_errors.append(f"Duplicate staff_number in file: {staff_number}")
            elif staff_number in existing_numbers:
                row_errors.append(f"Staff number {staff_number} already exists")
            else:
                seen_numbers.add(staff_number)

        gender = str(row_data.get("gender", "") or "").strip().lower()
        if gender and gender not in ("male", "female"):
            row_errors.append(f"gender must be Male or Female, got: {gender!r}")
        gender = gender or None

        status = str(row_data.get("status", "") or "").strip().lower() or "active"
        valid_statuses = ("active", "on_leave", "transferred", "retired")
        if status not in valid_statuses:
            row_errors.append(f"status must be one of {valid_statuses}, got: {status!r}")
            status = "active"

        dob        = parse_date(row_data.get("date_of_birth"), "date_of_birth", row_errors)
        date_joined = parse_date(row_data.get("date_joined"), "date_joined", row_errors)

        if row_errors:
            skipped += 1
            errors.append({
                "row":  row_idx,
                "data": {"staff_number": staff_number or "", "name": f"{first_name} {last_name}"},
                "errors": row_errors,
            })
            continue

        db.add(Staff(
            school_id      = school.id,
            staff_number   = staff_number,
            first_name     = first_name,
            middle_name    = str(row_data.get("middle_name", "") or "").strip() or None,
            last_name      = last_name,
            title          = str(row_data.get("title", "") or "").strip() or None,
            gender         = gender,
            date_of_birth  = dob,
            phone          = str(row_data.get("phone", "") or "").strip() or None,
            date_joined    = date_joined,
            license_number = str(row_data.get("license_number", "") or "").strip() or None,
            specialization = str(row_data.get("specialization", "") or "").strip() or None,
            status         = status,
        ))
        if staff_number:
            existing_numbers.add(staff_number)
        imported += 1

    await db.commit()
    return {
        "imported": imported,
        "skipped":  skipped,
        "errors":   errors,
        "message": (
            f"Successfully imported {imported} staff member{'s' if imported != 1 else ''}"
            + (f". {skipped} row{'s' if skipped != 1 else ''} skipped." if skipped else ".")
        ),
    }


# ── Qualifications ────────────────────────────────────────────────────────
@router.post("/{staff_id}/qualifications", response_model=QualificationResponse, status_code=201)
async def add_qualification(
    staff_id: UUID,
    body: QualificationCreate,
    db: DB,
    current_user: CurrentUser,
    school: CurrentSchool,
):
    if not _can_manage(current_user, staff_id):
        raise HTTPException(status_code=403, detail="Not authorised")

    staff = await db.get(Staff, staff_id)
    if not staff or staff.school_id != school.id:
        raise HTTPException(status_code=404, detail="Staff member not found")

    qual = StaffQualification(staff_id=staff_id, **body.model_dump())
    db.add(qual)
    await db.commit()
    await db.refresh(qual)
    return qual


@router.delete("/{staff_id}/qualifications/{qual_id}", status_code=204)
async def remove_qualification(
    staff_id: UUID,
    qual_id: UUID,
    db: DB,
    current_user: CurrentUser,
    school: CurrentSchool,
):
    if not _can_manage(current_user, staff_id):
        raise HTTPException(status_code=403, detail="Not authorised")

    staff = await db.get(Staff, staff_id)
    if not staff or staff.school_id != school.id:
        raise HTTPException(status_code=404, detail="Staff member not found")

    result = await db.execute(
        select(StaffQualification).where(
            StaffQualification.id == qual_id,
            StaffQualification.staff_id == staff_id,
        )
    )
    qual = result.scalar_one_or_none()
    if not qual:
        raise HTTPException(status_code=404, detail="Qualification not found")
    await db.delete(qual)
    await db.commit()


# ── Promotions ────────────────────────────────────────────────────────────
@router.post("/{staff_id}/promotions", response_model=PromotionResponse, status_code=201)
async def add_promotion(
    staff_id: UUID,
    body: PromotionCreate,
    db: DB,
    current_user: CurrentUser,
    school: CurrentSchool,
):
    if not _can_manage(current_user, staff_id):
        raise HTTPException(status_code=403, detail="Not authorised")

    staff = await db.get(Staff, staff_id)
    if not staff or staff.school_id != school.id:
        raise HTTPException(status_code=404, detail="Staff member not found")

    promo = StaffPromotion(staff_id=staff_id, **body.model_dump())
    db.add(promo)
    await db.commit()
    await db.refresh(promo)
    return promo


@router.delete("/{staff_id}/promotions/{promo_id}", status_code=204)
async def remove_promotion(
    staff_id: UUID,
    promo_id: UUID,
    db: DB,
    _: Annotated[User, Depends(require_roles("school_admin", "headteacher"))],
    school: CurrentSchool,
):
    staff = await db.get(Staff, staff_id)
    if not staff or staff.school_id != school.id:
        raise HTTPException(status_code=404, detail="Staff member not found")

    result = await db.execute(
        select(StaffPromotion).where(
            StaffPromotion.id == promo_id,
            StaffPromotion.staff_id == staff_id,
        )
    )
    promo = result.scalar_one_or_none()
    if not promo:
        raise HTTPException(status_code=404, detail="Promotion record not found")
    await db.delete(promo)
    await db.commit()
