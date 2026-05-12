from typing import Annotated, List, Optional
from datetime import date as date_type
from io import BytesIO
from uuid import UUID

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers as xl_numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import selectinload
from sqlalchemy import select
from pwdlib import PasswordHash
from pwdlib.hashers.argon2 import Argon2Hasher

from app.dependencies import CurrentUser, CurrentSchool, DB, require_roles
from app.models.user import User
from app.models.student import Student
from app.models.student_contact import StudentContact
from app.models.academic import Class as ClassModel, AcademicYear
from app.models.enrollment import Enrollment
from app.models.school_house import SchoolHouse
from app.schemas.student import (
    StudentCreate, StudentUpdate,
    StudentResponse, StudentContactCreate, StudentContactResponse
)

router = APIRouter()


# ── GET /api/students ──────────────────────────────────────────────────────

@router.get("/", response_model=List[StudentResponse])
async def list_students(
    user: CurrentUser,
    school: CurrentSchool,
    db: DB,
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, le=5000),
):
    query = select(Student).options(selectinload(Student.contacts)).where(Student.school_id == school.id)

    if search:
        query = query.where(
            Student.first_name.ilike(f"%{search}%") |
            Student.last_name.ilike(f"%{search}%") |
            Student.student_number.ilike(f"%{search}%")
        )
    if status:
        query = query.where(Student.status == status)

    query = query.offset(skip).limit(limit)
    result = await db.execute(query)
    return result.scalars().all()


# ── GET /api/students/{id} ─────────────────────────────────────────────────

@router.get("/{student_id}", response_model=StudentResponse)
async def get_student(
    student_id: UUID,
    user: CurrentUser,
    school: CurrentSchool,
    db: DB,
):
    result = await db.execute(
        select(Student)
        .options(selectinload(Student.contacts))
        .where(
            Student.id == student_id,
            Student.school_id == school.id,
        )
    )

    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student


# ── POST /api/students ─────────────────────────────────────────────────────

@router.post("/", response_model=StudentResponse, status_code=201)
async def create_student(
    body: StudentCreate,
    user: Annotated[User, Depends(require_roles(
        "school_admin", "headteacher"
    ))],
    school: CurrentSchool,
    db: DB,
):
    # Check student number not already used in this school
    exists = await db.execute(
        select(Student).where(
            Student.school_id == school.id,
            Student.student_number == body.student_number,
        )
    )
    if exists.scalar_one_or_none():
        raise HTTPException(
            status_code=409,
            detail=f"Student number '{body.student_number}' already exists",
        )

    student = Student(
        school_id=school.id,
        student_number=body.student_number,
        first_name=body.first_name,
        middle_name=body.middle_name,
        last_name=body.last_name,
        date_of_birth=body.date_of_birth,
        gender=body.gender,
        photo_url=body.photo_url,
        home_address=body.home_address,
        admission_date=body.admission_date,
        house=body.house,
        programme=body.programme,
        notes=body.notes,
    )
    db.add(student)
    await db.flush()   # get student.id before adding contacts

    # Add contacts if provided
    for c in body.contacts:
        contact = StudentContact(
            school_id=school.id,
            student_id=student.id,
            **c.model_dump(),
        )
        db.add(contact)

    await db.commit()
    result = await db.execute(
        select(Student)
        .options(selectinload(Student.contacts))
        .where(Student.id == student.id)
    )
    return result.scalar_one()





# ── PATCH /api/students/{id} ───────────────────────────────────────────────

@router.patch("/{student_id}", response_model=StudentResponse)
async def update_student(
    student_id: UUID,
    body: StudentUpdate,
    user: Annotated[User, Depends(require_roles("school_admin", "headteacher"))],
    school: CurrentSchool,
    db: DB,
):
    result = await db.execute(
        select(Student).where(
            Student.id == student_id,
            Student.school_id == school.id,
        )
    )
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    # Only update fields that were actually sent
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(student, field, value)

    await db.commit()
    result = await db.execute(
        select(Student)
        .options(selectinload(Student.contacts))
        .where(Student.id == student.id)
    )
    return result.scalar_one()


# ── DELETE /api/students/{id} ──────────────────────────────────────────────

@router.delete("/{student_id}", status_code=204)
async def delete_student(
    student_id: UUID,
    user: Annotated[CurrentUser, Depends(require_roles("school_admin"))],
    school: CurrentSchool,
    db: DB,
):
    result = await db.execute(
        select(Student).where(
            Student.id == student_id,
            Student.school_id == school.id,
        )
    )
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    await db.delete(student)
    await db.commit()


# ── POST /api/students/{id}/contacts ──────────────────────────────────────

@router.post("/{student_id}/contacts",
             response_model=StudentContactResponse,
             status_code=201)
async def add_contact(
    student_id: UUID,
    body: StudentContactCreate,
    user: Annotated[User, Depends(require_roles("school_admin", "headteacher"))],
    school: CurrentSchool,
    db: DB,
):
    # Verify student belongs to this school
    result = await db.execute(
        select(Student).where(
            Student.id == student_id,
            Student.school_id == school.id,
        )
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Student not found")

    # If this is primary contact, unset any existing primary
    if body.is_primary_contact:
        existing = await db.execute(
            select(StudentContact).where(
                StudentContact.student_id == student_id,
                StudentContact.is_primary_contact == True,
            )
        )
        for c in existing.scalars().all():
            c.is_primary_contact = False

    contact = StudentContact(
        school_id=school.id,
        student_id=student_id,
        **body.model_dump(),
    )
    db.add(contact)
    await db.commit()
    await db.refresh(contact)
    return contact
# ══════════════════════════════════════════════════════════════════════════
# BULK UPLOAD
# ══════════════════════════════════════════════════════════════════════════


@router.get("/upload/template")
async def download_template(
    user: CurrentUser,
    school: CurrentSchool,
    db: DB,
):
    """
    Download school-branded Excel template for bulk student upload.
    - Header color matches school accent color
    - Class dropdown from school's active classes
    - House dropdown if school has houses configured
    - Phone columns formatted as Text (preserves leading zero)
    - Relation dropdown with capitalized display names
    """
    # ── Fetch school data ─────────────────────────────────────────
    classes_result = await db.execute(
        select(ClassModel).where(
            ClassModel.school_id == school.id,
            ClassModel.is_active == True,
        ).order_by(ClassModel.level_group, ClassModel.level_number, ClassModel.stream)
    )
    classes = classes_result.scalars().all()

    houses_result = await db.execute(
        select(SchoolHouse).where(
            SchoolHouse.school_id == school.id,
            SchoolHouse.is_active == True,
        ).order_by(SchoolHouse.order)
    )
    houses = houses_result.scalars().all()

    year_result = await db.execute(
        select(AcademicYear).where(
            AcademicYear.school_id == school.id,
            AcademicYear.is_current == True,
        )
    )
    current_year = year_result.scalar_one_or_none()

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════
    # HIDDEN SHEET: _data — stores lookup values for dropdowns
    # ══════════════════════════════════════════════════════════════
    ws_data = wb.create_sheet("_data")
    ws_data.sheet_state = "hidden"

    # Class names + IDs
    ws_data["A1"] = "class_name"
    ws_data["B1"] = "class_id"
    for i, cls in enumerate(classes, 2):
        ws_data.cell(row=i, column=1, value=cls.name)
        ws_data.cell(row=i, column=2, value=str(cls.id))

    # House names
    ws_data["C1"] = "house_name"
    for i, house in enumerate(houses, 2):
        ws_data.cell(row=i, column=3, value=house.name)

    num_classes = len(classes)
    num_houses  = len(houses)

    # ══════════════════════════════════════════════════════════════
    # MAIN SHEET: Students
    # ══════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Students"

    # ── School brand color ────────────────────────────────────────
    school_color  = (school.accent_color or "#1a6b3c").lstrip("#").upper()
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", fgColor=school_color)
    required_fill = PatternFill("solid", fgColor="E8F5E9")
    optional_fill = PatternFill("solid", fgColor="F5F5F5")
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Title row ─────────────────────────────────────────────────
    ws.merge_cells("A1:O1")
    year_label = current_year.name if current_year else "No current year set"
    ws["A1"] = f"{school.name}  |  Student Import  |  {year_label}"
    ws["A1"].font = Font(bold=True, size=13, color=school_color)
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 30

    # ── Instructions ──────────────────────────────────────────────
    ws.merge_cells("A2:O2")
    class_note = (
        f"{num_classes} class(es) available in dropdown."
        if num_classes else
        "No classes found — create classes first."
    )
    house_note = (
        f"{num_houses} house(s) in dropdown."
        if num_houses else
        "No houses configured — type house name freely."
    )
    ws["A2"] = (
        f"Fill from row 5. Green = required. "
        f"Date: YYYY-MM-DD. Phone: include leading zero e.g. 0244123456. "
        f"{class_note}  {house_note}"
    )
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["A2"].alignment = center_align
    ws.row_dimensions[2].height = 25

    # ── Column definitions ────────────────────────────────────────
    # (header, width, required, example, is_phone)
    columns = [
        ("student_number",     15, True,  "SHS001",      False),
        ("first_name",         15, True,  "Kofi",         False),
        ("middle_name",        15, False, "Adu",          False),
        ("last_name",          15, True,  "Mensah",       False),
        ("gender",             10, True,  "Male",         False),
        ("date_of_birth",      15, False, "2008-03-15",   False),
        ("admission_date",     15, False, "2024-09-01",   False),
        ("house",              15, False, houses[0].name if houses else "Unity", False),
        ("contact_first_name", 18, False, "Emmanuel",     False),
        ("contact_last_name",  18, False, "Mensah",       False),
        ("contact_relation",   18, False, "Father",       False),
        ("contact_phone",      15, False, "0244123456",   True),
        ("contact_phone2",     15, False, "0201234567",   True),
        ("contact_is_parent",  15, False, "Yes",          False),
        ("class",              22, False, classes[0].name if classes else "", False),
    ]

    # ── Header row (row 3) ────────────────────────────────────────
    for col_idx, (header, width, required, _, _ph) in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 22

    # ── Example row (row 4) ───────────────────────────────────────
    for col_idx, (_, _, required, example, _ph) in enumerate(columns, 1):
        cell = ws.cell(row=4, column=col_idx, value=example)
        cell.fill = required_fill if required else optional_fill
        cell.font = Font(italic=True, color="888888", size=10)

    # ── Format phone columns as Text (preserves leading zero) ─────
    text_format = xl_numbers.FORMAT_TEXT
    for col_idx, (_, _, _, _, is_phone) in enumerate(columns, 1):
        if is_phone:
            col_letter = get_column_letter(col_idx)
            for row in range(4, 10001):
                ws[f"{col_letter}{row}"].number_format = text_format

    # ── Data validations ──────────────────────────────────────────

    # Gender — capitalized display
    gender_dv = DataValidation(
        type="list",
        formula1='"Male,Female"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid gender",
        error="Select Male or Female",
    )
    ws.add_data_validation(gender_dv)
    gender_dv.sqref = "E5:E10000"

    # Contact is_parent
    parent_dv = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True,
    )
    ws.add_data_validation(parent_dv)
    parent_dv.sqref = "N5:N10000"

    # Relation — capitalized display
    relation_dv = DataValidation(
        type="list",
        formula1='"Father,Mother,Grandfather,Grandmother,Uncle,Aunt,Brother,Sister,Guardian,Other"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid relation",
        error="Select a relation from the list",
    )
    ws.add_data_validation(relation_dv)
    relation_dv.sqref = "K5:K10000"

    # House dropdown (only if school has houses)
    if num_houses > 0:
        house_range = f"_data!$C$2:$C${num_houses + 1}"
        house_dv = DataValidation(
            type="list",
            formula1=house_range,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid house",
            error="Select a house from the list",
        )
        ws.add_data_validation(house_dv)
        house_dv.sqref = "H5:H10000"

    # Class dropdown (from hidden sheet)
    if num_classes > 0:
        class_range = f"_data!$A$2:$A${num_classes + 1}"
        class_dv = DataValidation(
            type="list",
            formula1=class_range,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid class",
            error="Select a class from the list",
        )
        ws.add_data_validation(class_dv)
        class_dv.sqref = "O5:O10000"

    # ── Freeze header rows ────────────────────────────────────────
    ws.freeze_panes = "A5"

    # ── Save ──────────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{school.slug}_student_import.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/upload")
async def bulk_upload_students(
    user: Annotated[User, Depends(require_roles("school_admin", "headteacher"))],
    school: CurrentSchool,
    db: DB,
    file: UploadFile = File(...),
    class_id: Optional[UUID] = None,
    academic_year_id: Optional[UUID] = None,
    start_date: Optional[date_type] = None,
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "File must be an Excel file (.xlsx or .xls)")

    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(400, "File must be under 10 MB")
    try:
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "Could not read Excel file. Please use the template.")

    HEADER_ROW = 3
    DATA_START  = 5

    headers = []
    for cell in ws[HEADER_ROW]:
        if cell.value:
            headers.append(str(cell.value).strip())

    required_headers = {"student_number", "first_name", "last_name"}
    missing = required_headers - set(headers)
    if missing:
        raise HTTPException(
            400,
            f"Missing required columns: {missing}. Please use the downloaded template."
        )

    class_ = None
    if class_id:
        class_result = await db.execute(
            select(ClassModel).where(
                ClassModel.id == class_id,
                ClassModel.school_id == school.id,
            )
        )
        class_ = class_result.scalar_one_or_none()
        if not class_:
            raise HTTPException(404, "Class not found")
        if not academic_year_id or not start_date:
            raise HTTPException(
                400,
                "academic_year_id and start_date are required when class_id is provided"
            )

    imported  = 0
    skipped   = 0
    errors    = []
    seen_numbers = set()

    existing_result = await db.execute(
        select(Student.student_number).where(Student.school_id == school.id)
    )
    existing_numbers = {row[0] for row in existing_result.all()}

    def parse_date(val, field_name: str, row_errors: list):
        if not val:
            return None
        try:
            if isinstance(val, date_type):
                return val
            return date_type.fromisoformat(str(val).strip())
        except ValueError:
            row_errors.append(f"{field_name} must be YYYY-MM-DD, got: {val}")
            return None

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=DATA_START, values_only=True), start=DATA_START
    ):
        if not any(row):
            continue

        row_data = {}
        for i, header in enumerate(headers):
            row_data[header] = row[i] if i < len(row) else None

        row_errors = []

        student_number = str(row_data.get("student_number", "") or "").strip()
        first_name     = str(row_data.get("first_name", "") or "").strip()
        last_name      = str(row_data.get("last_name", "") or "").strip()

        if not student_number:
            row_errors.append("student_number is required")
        if not first_name:
            row_errors.append("first_name is required")
        if not last_name:
            row_errors.append("last_name is required")

        if student_number:
            if student_number in seen_numbers:
                row_errors.append(f"Duplicate student_number in file: {student_number}")
            elif student_number in existing_numbers:
                row_errors.append(f"Student number {student_number} already exists")
            else:
                seen_numbers.add(student_number)

        # Normalize to lowercase — template shows Male/Female
        gender = str(row_data.get("gender", "") or "").strip().lower()
        if gender and gender not in ("male", "female"):
            row_errors.append(f"gender must be Male or Female, got: {gender}")
        gender = gender or None

        dob          = parse_date(row_data.get("date_of_birth"),  "date_of_birth", row_errors)
        admission_dt = parse_date(row_data.get("admission_date"), "admission_date", row_errors)

        if row_errors:
            skipped += 1
            errors.append({
                "row":    row_idx,
                "data":   {"student_number": student_number,
                           "name": f"{first_name} {last_name}"},
                "errors": row_errors,
            })
            continue

        student = Student(
            school_id      = school.id,
            student_number = student_number,
            first_name     = first_name,
            middle_name    = str(row_data.get("middle_name", "") or "").strip() or None,
            last_name      = last_name,
            gender         = gender,
            date_of_birth  = dob,
            admission_date = admission_dt,
            house          = str(row_data.get("house", "") or "").strip() or None,
        )
        db.add(student)
        await db.flush()

        contact_first = str(row_data.get("contact_first_name", "") or "").strip()
        contact_phone = str(row_data.get("contact_phone", "") or "").strip()

        if contact_first and contact_phone:
            is_parent_val = str(
                row_data.get("contact_is_parent", "Yes") or "Yes"
            ).strip().lower()
            db.add(StudentContact(
                school_id          = school.id,
                student_id         = student.id,
                first_name         = contact_first,
                last_name          = str(row_data.get("contact_last_name", "") or "").strip() or None,
                relation           = str(row_data.get("contact_relation", "guardian") or "guardian").strip().lower(),
                phone              = contact_phone,
                phone2             = str(row_data.get("contact_phone2", "") or "").strip() or None,
                is_parent          = is_parent_val == "yes",
                is_primary_contact = True,
                can_pickup         = True,
                receives_sms       = True,
                is_alive           = True,
            ))

        if class_id and academic_year_id and start_date:
            db.add(Enrollment(
                school_id        = school.id,
                student_id       = student.id,
                class_id         = class_id,
                academic_year_id = academic_year_id,
                start_date       = start_date,
                status           = "active",
            ))

        imported += 1
        existing_numbers.add(student_number)

    await db.commit()

    return {
        "imported": imported,
        "skipped":  skipped,
        "errors":   errors,
        "message":  (
            f"Successfully imported {imported} students"
            + (f" and enrolled in {class_.name}" if class_ else "")
            + (f". {skipped} rows skipped." if skipped else ".")
        ),
    }


# ══════════════════════════════════════════════════════════════════════════
# STUDENT PORTAL CREDENTIALS
# ══════════════════════════════════════════════════════════════════════════

@router.post("/{student_id}/enable-portal")
async def enable_student_portal(
    student_id: UUID,
    user: CurrentUser,
    school: CurrentSchool,
    db: DB,
):
    """
    Enable student portal access.
    Sets default PIN = date of birth (DDMMYYYY).
    Falls back to student_number if no DOB.
    Student must change PIN on first login.
    """
    if user.role not in ("school_admin", "headteacher"):
        raise HTTPException(403, "Only admin or headteacher can enable portal access")

    result = await db.execute(
        select(Student).where(
            Student.id == student_id,
            Student.school_id == school.id,
        )
    )
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(404, "Student not found")

    # Check school has portal enabled for this student's level
    year_res = await db.execute(
        select(AcademicYear).where(
            AcademicYear.school_id == school.id,
            AcademicYear.is_current == True,
        )
    )
    year = year_res.scalar_one_or_none()

    if year:
        enrollment_res = await db.execute(
            select(Enrollment).where(
                Enrollment.student_id == student_id,
                Enrollment.academic_year_id == year.id,
                Enrollment.status == "active",
            )
        )
        enrollment = enrollment_res.scalar_one_or_none()
        if enrollment:
            class_res = await db.execute(
                select(ClassModel).where(ClassModel.id == enrollment.class_id)
            )
            cls = class_res.scalar_one_or_none()
            if cls and not school.portal_enabled_for_level(
                cls.level_group, cls.level_number
            ):
                raise HTTPException(
                    400,
                    f"Student portal is not enabled for {cls.level_group.upper()} level. "
                    f"Enable it in school settings first."
                )

    pwd = PasswordHash((Argon2Hasher(),))
    if student.date_of_birth:
        default_pin = student.date_of_birth.strftime("%d%m%Y")
    else:
        default_pin = student.student_number

    student.pin_hash = pwd.hash(default_pin)
    await db.commit()

    return {
        "message": "Portal access enabled",
        "student_number": student.student_number,
        "default_pin": default_pin,
        "note": "Student must change PIN on first login",
    }


@router.post("/{student_id}/disable-portal")
async def disable_student_portal(
    student_id: UUID,
    user: CurrentUser,
    school: CurrentSchool,
    db: DB,
):
    """Disable student portal access."""
    if user.role not in ("school_admin", "headteacher"):
        raise HTTPException(403, "Only admin or headteacher can disable portal access")

    result = await db.execute(
        select(Student).where(
            Student.id == student_id,
            Student.school_id == school.id,
        )
    )
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(404, "Student not found")

    student.pin_hash = None
    await db.commit()
    return {"message": "Portal access disabled"}


@router.post("/{student_id}/reset-pin")
async def reset_student_pin(
    student_id: UUID,
    user: CurrentUser,
    school: CurrentSchool,
    db: DB,
):
    """
    Reset student PIN to default (date of birth or student number).
    Used when student forgets PIN.
    """
    if user.role not in ("school_admin", "headteacher", "teacher"):
        raise HTTPException(403, "Not authorised to reset PIN")

    result = await db.execute(
        select(Student).where(
            Student.id == student_id,
            Student.school_id == school.id,
        )
    )
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(404, "Student not found")

    if not student.pin_hash:
        raise HTTPException(400, "Portal not enabled for this student")

    pwd = PasswordHash((Argon2Hasher(),))
    if student.date_of_birth:
        default_pin = student.date_of_birth.strftime("%d%m%Y")
    else:
        default_pin = student.student_number

    student.pin_hash = pwd.hash(default_pin)
    await db.commit()

    return {
        "message": "PIN reset to default",
        "default_pin": default_pin,
        "note": "Student must change PIN on next login",
    }


@router.post("/bulk-enable-portal")
async def bulk_enable_portal(
    class_id: UUID,
    user: CurrentUser,
    school: CurrentSchool,
    db: DB,
):
    """
    Enable portal for all active students in a class at once.
    Useful at start of term.
    """
    if user.role not in ("school_admin", "headteacher"):
        raise HTTPException(403, "Only admin or headteacher can bulk enable portal")

    year_res = await db.execute(
        select(AcademicYear).where(
            AcademicYear.school_id == school.id,
            AcademicYear.is_current == True,
        )
    )
    year = year_res.scalar_one_or_none()
    if not year:
        raise HTTPException(404, "No current academic year")

    enrollments_res = await db.execute(
        select(Enrollment)
        .options(selectinload(Enrollment.student))
        .where(
            Enrollment.school_id == school.id,
            Enrollment.class_id == class_id,
            Enrollment.academic_year_id == year.id,
            Enrollment.status == "active",
        )
    )
    enrollments = enrollments_res.scalars().all()

    pwd = PasswordHash((Argon2Hasher(),))
    enabled = 0
    skipped = 0

    for enrollment in enrollments:
        student = enrollment.student
        if student.pin_hash:
            skipped += 1
            continue
        if student.date_of_birth:
            default_pin = student.date_of_birth.strftime("%d%m%Y")
        else:
            default_pin = student.student_number
        student.pin_hash = pwd.hash(default_pin)
        enabled += 1

    await db.commit()
    return {
        "enabled": enabled,
        "skipped": skipped,
        "message": f"Portal enabled for {enabled} students. "
                   f"{skipped} already had access.",
    }


# ══════════════════════════════════════════════════════════════════════════
# TRANSFER INBOUND
# ══════════════════════════════════════════════════════════════════════════

@router.post("/{student_id}/transfer-in")
async def transfer_student_in(
    student_id: UUID,
    user: CurrentUser,
    school: CurrentSchool,
    db: DB,
    class_id: Optional[UUID] = None,
    academic_year_id: Optional[UUID] = None,
    start_date: Optional[date_type] = None,
    previous_school: Optional[str] = None,
    notes: Optional[str] = None,
):
    """
    Record a student transferring INTO this school.
    Creates enrollment in the specified class.
    Updates student status to active.
    """
    result = await db.execute(
        select(Student).where(
            Student.id == student_id,
            Student.school_id == school.id,
        )
    )
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(404, "Student not found")

    if student.status == "active":
        raise HTTPException(400, "Student is already active")

    if not class_id or not academic_year_id or not start_date:
        raise HTTPException(
            400,
            "class_id, academic_year_id and start_date are required"
        )

    # Check not already enrolled this year
    existing = await db.execute(
        select(Enrollment).where(
            Enrollment.student_id == student_id,
            Enrollment.academic_year_id == academic_year_id,
            Enrollment.status == "active",
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(409, "Student already has an active enrollment this year")

    student.status = "active"

    enrollment_notes = f"Transfer in from {previous_school}. " if previous_school else "Transfer in. "
    if notes:
        enrollment_notes += notes

    db.add(Enrollment(
        school_id=school.id,
        student_id=student_id,
        class_id=class_id,
        academic_year_id=academic_year_id,
        start_date=start_date,
        status="active",
        notes=enrollment_notes,
    ))

    await db.commit()
    return {
        "message": "Student transfer recorded successfully",
        "student_number": student.student_number,
        "previous_school": previous_school,
    }
